"""
darth-bot/inventory.py
======================
Reads inventory data from the Destiny Voyager toolkit so the bot can answer
"good build with my current weapons" — the killer feature no other
Destiny bot can offer.

Two sources:
  1. user_config.json  — has saved loadouts, build_focus, item_tags
  2. my_loadouts.xlsx  — has INVENTORY + MY LOADOUTS sheets populated by
                          destiny-voyager/fetch_inventory.py (v0.2.0 feature)

Returns a compact text summary suitable for stuffing into the LLM context.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Optional

from .config import DESTINY_VOYAGER_CONFIG, DESTINY_VOYAGER_WORKBOOK


def _read_config() -> dict:
    if not DESTINY_VOYAGER_CONFIG.exists():
        return {}
    try:
        return json.loads(DESTINY_VOYAGER_CONFIG.read_text())
    except Exception:
        return {}


def _read_workbook_inventory() -> list[dict]:
    """Extract item rows from the INVENTORY sheet, if present."""
    if not DESTINY_VOYAGER_WORKBOOK.exists():
        return []
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []
    try:
        wb = load_workbook(DESTINY_VOYAGER_WORKBOOK, read_only=True)
        if "INVENTORY" not in wb.sheetnames:
            return []
        ws = wb["INVENTORY"]
        items = []
        # Columns: # Name Tier Element Type Slot Power Tag Location InstanceID
        for row in ws.iter_rows(min_row=4, values_only=True):
            if not row or not row[1]:
                continue
            items.append({
                "name": row[1],
                "tier": row[2] or "",
                "element": row[3] or "",
                "type": row[4] or "",
                "slot": row[5] or "",
                "power": row[6] or "",
                "tag": row[7] or "",
                "location": row[8] or "",
            })
        return items
    except Exception:
        return []


def build_context(focus: str = "all", max_items: int = 80) -> str:
    """
    Compose an inventory summary string for the LLM context window.

    focus: one of
        "all"          — top tagged + equipped items across the board
        "weapons"      — focus on weapons
        "armor"        — focus on armor
        "<class>"      — focus on one of hunter / titan / warlock
        "<activity>"   — pvp, pve, raid, dungeon (filters by tag/notes)

    Returns plaintext. Empty string if no inventory data is available.
    """
    cfg = _read_config()
    items = _read_workbook_inventory()

    if not items and not cfg.get("saved_loadouts"):
        return ""

    lines = []

    # Build focus summary if present
    bf = cfg.get("build_focus") or {}
    if bf:
        lines.append(
            f"User build focus: archetype={bf.get('archetype','?')}, "
            f"goals={', '.join(bf.get('goals',[]))}, "
            f"primary stats={', '.join(bf.get('target_stats',[]))}."
        )

    primary_class = cfg.get("primary_class")
    if primary_class:
        lines.append(f"Primary class: {primary_class}.")
    bungie_name = cfg.get("bungie_name")
    if bungie_name:
        lines.append(f"Guardian: {bungie_name}.")

    # Filter items by focus
    filtered = items
    f = focus.lower()
    if f in ("hunter", "titan", "warlock"):
        filtered = [i for i in items if f in (i.get("location", "") or "").lower()]
    elif f == "weapons":
        filtered = [i for i in items if i.get("slot", "") in ("Kinetic", "Energy", "Heavy")]
    elif f == "armor":
        filtered = [i for i in items if i.get("slot", "") in ("Helmet", "Gauntlets", "Chest Armor", "Leg Armor", "Class Armor")]

    # Prioritize: tagged favorite > tagged keep > equipped > everything else
    def rank(it):
        tag = (it.get("tag") or "").lower()
        loc = (it.get("location") or "").lower()
        if "equipped" in loc:        return 0
        if tag == "favorite":        return 1
        if tag == "keep":            return 2
        return 3
    filtered.sort(key=rank)
    filtered = filtered[:max_items]

    if filtered:
        lines.append("")
        lines.append(f"Inventory snapshot ({len(filtered)} items, filtered for: {focus}):")
        for it in filtered:
            bits = [
                f"{it['name']}",
                f"({it['tier']})",
                f"[{it['type']}]",
                f"{it['element']}",
                f"pow {it['power']}",
                f"@ {it['location']}",
            ]
            if it.get("tag"):
                bits.append(f"#tag:{it['tag']}")
            lines.append(" · ".join(b for b in bits if b and b != "()"))

    # Saved loadouts
    saved = cfg.get("saved_loadouts") or []
    if saved:
        lines.append("")
        lines.append(f"Saved loadouts ({len(saved)}):")
        for ld in saved[:20]:
            lines.append(f"  - {ld.get('class','?')} :: {ld.get('name','?')} ({ld.get('source','?')})")

    return "\n".join(lines).strip()


def has_inventory() -> bool:
    return DESTINY_VOYAGER_CONFIG.exists() and bool(_read_workbook_inventory())
