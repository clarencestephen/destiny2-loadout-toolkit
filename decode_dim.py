r"""
decode_dim.py
=============
Reads DIM share URLs from user_config.json, resolves every weapon / armor /
mod / aspect / fragment hash via the Bungie manifest, and writes a verbose
"DIM LOADOUTS (FULL)" tab to your workbook.

Usage:
    python3 decode_dim.py

Prerequisites:
    - user_config.json (run `python3 setup.py` first if you don't have one)
    - openpyxl installed (`pip install -r requirements.txt`)

The script:
1. Loads user_config.json
2. Downloads + caches the Bungie manifest (~150 MB JSON, cached for 7 days)
3. Parses each DIM URL → extracts equipped item hashes, mod hashes, subclass overrides
4. Resolves every hash → display name
5. Rewrites the "DIM LOADOUTS (FULL)" tab. Every other tab is untouched.
"""

import json
import sys
import time
import urllib.parse
import urllib.request
from pathlib import Path

CONFIG_PATH = Path("user_config.json")

# Position 0-2 = weapons (kinetic/energy/heavy), 3-7 = armor, 8 = subclass
SLOTS = ["Kinetic", "Energy", "Heavy", "Helmet", "Gauntlets",
         "Chest", "Legs", "Class Item", "Subclass"]

BUCKET_NAMES = {
    "14239492":   "Helmet mods",
    "20886954":   "Gauntlet mods",
    "1585787867": "Chest mods",
    "3448274439": "Leg mods",
    "3551918588": "Class item mods",
}

CLASS_COLORS = {"Hunter": "2563EB", "Titan": "DC2626", "Warlock": "7C3AED"}


def load_config():
    if not CONFIG_PATH.exists():
        sys.exit(f"ERROR: {CONFIG_PATH} not found. Run `python3 setup.py` first.")
    cfg = json.loads(CONFIG_PATH.read_text())
    if not cfg.get("api_key") or "PASTE" in cfg["api_key"]:
        sys.exit("ERROR: api_key missing/placeholder in user_config.json. "
                 "Run `python3 setup.py` or edit the file directly.")
    if not cfg.get("dim_loadouts"):
        sys.exit("ERROR: no DIM loadouts in user_config.json. "
                 "Run `python3 add_loadout.py` to add one.")
    return cfg


def bungie_get(url, api_key):
    req = urllib.request.Request(
        url,
        headers={"X-API-Key": api_key,
                 "User-Agent": "order-66/1.0"}
    )
    with urllib.request.urlopen(req, timeout=60) as r:
        return json.loads(r.read().decode("utf-8"))


def get_manifest_paths(api_key):
    data = bungie_get("https://www.bungie.net/Platform/Destiny2/Manifest/", api_key)
    return data["Response"]["jsonWorldComponentContentPaths"]["en"]


def download_manifest_table(table_name, manifest_paths, api_key, cache_dir):
    cache_dir.mkdir(exist_ok=True)
    cache_file = cache_dir / f"{table_name}.json"
    version_file = cache_dir / f"{table_name}.version"

    path = manifest_paths[table_name]
    if cache_file.exists() and version_file.exists():
        if version_file.read_text() == path:
            print(f"  Using cached {table_name}")
            return json.loads(cache_file.read_text())

    print(f"  Downloading {table_name} (this is the slow step)...")
    data = bungie_get(f"https://www.bungie.net{path}", api_key)
    cache_file.write_text(json.dumps(data))
    version_file.write_text(path)
    return data


def name_of(hash_str, items, fallback_tables=None):
    """Return display name for a hash, or '?(hash)' if not found."""
    h = str(hash_str)
    if h.startswith("-"):
        h = str((1 << 32) + int(h))

    entry = items.get(h)
    if not entry and fallback_tables:
        for table in fallback_tables.values():
            if h in table:
                entry = table[h]
                break
    if not entry:
        return f"?({hash_str})"
    return entry.get("displayProperties", {}).get("name", f"?({hash_str})")


def parse_dim_url(url):
    """Fetch a DIM share page → extract loadout JSON."""
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=30) as r:
        html = r.read().decode("utf-8")

    marker = "loadouts?loadout="
    idx = html.find(marker)
    if idx < 0:
        raise ValueError(f"Could not find loadout in {url}. "
                         "DIM may have changed its page format — open an issue on GitHub.")
    start = idx + len(marker)
    # Blob is wrapped in href="..."; terminator is closing quote.
    # Fall back to ')' (legacy format) if no quote found.
    end = html.find('"', start)
    if end < 0:
        end = html.find(")", start)
    if end < 0:
        raise ValueError(f"Could not find end of loadout blob in {url}")
    encoded = html[start:end].rstrip(")").rstrip('"')
    return json.loads(urllib.parse.unquote(encoded))


def decode_loadout(loadout_json, items):
    """Decode one loadout JSON into a readable structure."""
    equipped = []
    for i, eq in enumerate(loadout_json.get("equipped", [])):
        h = str(eq["hash"])
        equipped.append({
            "slot": SLOTS[i] if i < len(SLOTS) else f"Slot {i}",
            "hash": h,
            "name": name_of(h, items),
            "overrides": eq.get("socketOverrides", {}),
        })

    # Resolve subclass aspects/fragments/abilities
    if equipped and equipped[-1].get("overrides"):
        sub = equipped[-1]
        sub["decoded_overrides"] = {
            k: name_of(v, items) for k, v in sub["overrides"].items()
        }

    params = loadout_json.get("parameters", {})
    armor_mods = [name_of(h, items) for h in params.get("mods", [])]

    armor_artifice = {}
    for bucket, mod_list in params.get("modsByBucket", {}).items():
        armor_artifice[BUCKET_NAMES.get(bucket, bucket)] = [
            name_of(h, items) for h in mod_list
        ]

    return {
        "equipped": equipped,
        "armor_mods": armor_mods,
        "armor_artifice": armor_artifice,
    }


def write_workbook(workbook_path, decoded_loadouts):
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = load_workbook(workbook_path)
    tab = "DIM LOADOUTS (FULL)"
    if tab in wb.sheetnames:
        del wb[tab]
    ws = wb.create_sheet(tab)

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def setc(c, val, **k):
        x = ws[c]
        x.value = val
        x.font = Font(name="Arial", bold=k.get("bold", False),
                      size=k.get("size", 11), color=k.get("color", "000000"),
                      italic=k.get("italic", False))
        if k.get("fill"):
            x.fill = PatternFill("solid", fgColor=k["fill"])
        x.alignment = Alignment(horizontal=k.get("align", "left"),
                                vertical="center", wrap_text=True)
        x.border = border

    widths = [4, 14, 12, 36, 36, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1
    ws.merge_cells(f"A{r}:F{r}")
    setc(f"A{r}", "  DIM LOADOUTS (FULL) — auto-resolved from Bungie manifest",
         bold=True, size=13, color="FFFFFF", fill="1F2937")
    ws.row_dimensions[r].height = 26
    r += 2

    for cls, name, activity, url, decoded in decoded_loadouts:
        ws.merge_cells(f"A{r}:F{r}")
        color = CLASS_COLORS.get(cls, "1F2937")
        setc(f"A{r}", f"  {cls.upper()}  ◆  {name}  ◆  {url}",
             bold=True, color="FFFFFF", fill=color, size=12)
        ws.row_dimensions[r].height = 24
        r += 1

        setc(f"A{r}", "#", bold=True, fill="F3F4F6", align="center")
        setc(f"B{r}", "Slot", bold=True, fill="F3F4F6")
        setc(f"C{r}", "Hash", bold=True, fill="F3F4F6", align="center")
        setc(f"D{r}", "Item", bold=True, fill="F3F4F6")
        setc(f"E{r}", "Notes / Aspects-Fragments", bold=True, fill="F3F4F6")
        setc(f"F{r}", "", fill="F3F4F6")
        r += 1

        for i, eq in enumerate(decoded["equipped"], 1):
            setc(f"A{r}", i, align="center")
            setc(f"B{r}", eq["slot"])
            setc(f"C{r}", eq["hash"], align="center", size=9, color="666666")
            setc(f"D{r}", eq["name"], bold=True)
            notes = ""
            if eq.get("decoded_overrides"):
                notes = " | ".join(eq["decoded_overrides"].values())
            setc(f"E{r}", notes, size=10)
            setc(f"F{r}", "")
            r += 1

        if decoded.get("armor_mods"):
            ws.merge_cells(f"A{r}:F{r}")
            setc(f"A{r}", f"  Mods: {', '.join(decoded['armor_mods'])}",
                 italic=True, size=10, color="666666", fill="F9FAFB")
            r += 1

        if decoded.get("armor_artifice"):
            for bucket, mods in decoded["armor_artifice"].items():
                ws.merge_cells(f"A{r}:F{r}")
                setc(f"A{r}", f"  {bucket}: {', '.join(mods)}",
                     italic=True, size=10, color="666666", fill="F9FAFB")
                r += 1

        r += 1

    wb.save(workbook_path)
    print(f'Wrote tab "{tab}" to {workbook_path}')


def main():
    cfg = load_config()
    api_key = cfg["api_key"]
    workbook = Path(cfg["workbook_path"])
    cache_dir = Path(cfg.get("manifest_cache_dir", "./manifest_cache"))

    if not workbook.exists():
        sys.exit(f"ERROR: workbook not found at {workbook}. "
                 "Run `python3 init_workbook.py` to create one.")

    print("Step 1/3: Fetching manifest paths...")
    paths = get_manifest_paths(api_key)

    print("Step 2/3: Downloading manifest tables (cached after first run)...")
    items = download_manifest_table(
        "DestinyInventoryItemDefinition", paths, api_key, cache_dir
    )
    print(f"  Loaded {len(items):,} item definitions")

    print("Step 3/3: Parsing DIM URLs + writing workbook...")
    results = []
    for entry in cfg["dim_loadouts"]:
        cls = entry["class"]
        name = entry["name"]
        activity = entry.get("activity", "")
        url = entry["url"]
        print(f"  {cls}: {name}")
        try:
            loadout = parse_dim_url(url)
        except Exception as e:
            print(f"    SKIPPED ({e})")
            continue
        decoded = decode_loadout(loadout, items)
        results.append((cls, name, activity, url, decoded))

    if not results:
        sys.exit("ERROR: no loadouts decoded successfully.")

    write_workbook(workbook, results)
    print(f'Done. Open the workbook → "DIM LOADOUTS (FULL)" tab.')


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nInterrupted.")
        sys.exit(130)
    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)
