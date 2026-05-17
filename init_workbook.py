"""
init_workbook.py
================
Builds a fresh, empty Destiny 2 loadout workbook at the path given on the
command line (or from user_config.json if no arg is provided).

Called automatically by setup.py during first-run install. Can also be
re-run standalone to regenerate a fresh template — but be warned, that
overwrites any existing file at that path.

Sheets created:
  1. START HERE          — onboarding instructions
  2. PRIORITIES          — exotic chase list, campaigns, farming routes
  3. WISHLIST            — DIM-CSV-compatible columns (paste your own here)
  4. HUNTER BUILDS       — empty class build template
  5. TITAN BUILDS        — empty class build template
  6. WARLOCK BUILDS      — empty class build template
  7. MOD REFERENCE       — artifact perk grid + armor mods reference
  8. DIM LOADOUTS (FULL) — auto-populated by decode_dim.py
"""

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CLASS_COLORS = {"Hunter": "2563EB", "Titan": "DC2626", "Warlock": "7C3AED"}

# Seed examples — populate sheets so they aren't blank on first open.
# Users replace these with their own builds. Format mirrors the column headers:
# (Build Name, Subclass, Exotic Armor, Kinetic, Energy, Heavy, Notes/Activity)
CLASS_EXAMPLES = {
    "Hunter": [
        ("Example: Still Hunt / Nighthawk", "Solar (Gunslinger)",
         "Celestial Nighthawk", "—", "Still Hunt", "Heavy of choice",
         "Boss DPS — Still Hunt + Golden Gun synergy"),
        ("Example: Lucky Pants Hand Cannon", "Strand (Threadrunner)",
         "Lucky Pants", "Malfeasance (or any HC)", "—", "Heavy of choice",
         "Hand cannon DPS / boss melt"),
        ("Example: Void Invis Crossbow", "Void (Nightstalker)",
         "Gyrfalcon's Hauberk", "—", "Wish-Keeper", "Heavy",
         "Volatile rounds, invis spam, ad clear"),
    ],
    "Titan": [
        ("Example: Wolfsbane Strand", "Strand (Berserker)",
         "Wishful Ignorance / Synthoceps", "—", "Wolfsbane", "Heavy",
         "Suspend + melee — high survivability"),
        ("Example: Sniper / Whirling Maelstrom", "Strand (Berserker)",
         "Synthoceps or Pyrogale Gauntlets", "—", "Sniper", "Heavy",
         "Whirling Maelstrom DPS"),
        ("Example: Bonk Hammer Solar", "Solar (Sunbreaker)",
         "Synthoceps", "Primary of choice", "Polaris Lance", "Sword",
         "Throwing hammer ad clear + restoration"),
    ],
    "Warlock": [
        ("Example: Heavy Sniper Prismatic", "Prismatic",
         "Getaway Artist", "Hand cannon", "Sniper of choice", "Heavy",
         "Boss DPS — Arc soul + bleak watcher freeze"),
        ("Example: Finality / Mint Stasis", "Stasis (Shadebinder)",
         "Osmiomancy Gloves", "Mint loadout primary", "Finality's Auger",
         "Heavy", "Linear fusion + turret freeze build"),
        ("Example: Song of Flame Solar", "Solar (Dawnblade)",
         "Sunbracers", "Primary", "Polaris Lance / Solar SMG", "Solar heavy",
         "Solar grenade spam + ignite chains"),
    ],
}


def setc(ws, ref, val, *, bold=False, size=11, color="000000",
         fill=None, italic=False, align="left", wrap=True, border=True):
    """Write a cell with consistent formatting."""
    c = ws[ref]
    c.value = val
    c.font = Font(name="Arial", bold=bold, size=size, color=color, italic=italic)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if border:
        c.border = BORDER


def section_header(ws, row, text, color="1F2937", height=24, cols=6):
    end = get_column_letter(cols)
    ws.merge_cells(f"A{row}:{end}{row}")
    setc(ws, f"A{row}", f"  {text}", bold=True, size=13, color="FFFFFF",
         fill=color, align="left")
    ws.row_dimensions[row].height = height


def build_start_here(wb):
    ws = wb.create_sheet("START HERE", 0)
    widths = [4, 24, 36, 36, 36, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    section_header(ws, 1, "DESTINY 2 LOADOUT TOOLKIT", color="111827", height=32)
    ws["A1"].font = Font(name="Arial", bold=True, size=18, color="FFFFFF")

    r = 3
    setc(ws, f"A{r}", "  Welcome. This workbook is your personal Destiny 2 planner.",
         bold=True, size=12, fill="F3F4F6")
    ws.merge_cells(f"A{r}:F{r}")
    ws.row_dimensions[r].height = 22
    r += 2

    blurb = [
        ("1. PRIORITIES",
         "Your exotic chase list, campaign queue, and farming routes. Edit row-by-row."),
        ("2. WISHLIST",
         "Weapons + armor you're hunting. DIM-CSV-compatible columns — you can paste rows from a DIM export."),
        ("3. HUNTER / TITAN / WARLOCK BUILDS",
         "One sheet per class. Add a build per row: super, exotic armor, weapons, mods, role."),
        ("4. MOD REFERENCE",
         "Current artifact perk grid + your armor mod loadout. Update each season."),
        ("5. DIM LOADOUTS (FULL)",
         "AUTO-GENERATED. Run `python3 decode_dim.py` to populate. Don't edit by hand — it's overwritten on each run."),
    ]
    for title, desc in blurb:
        setc(ws, f"A{r}", "", fill="F9FAFB")
        setc(ws, f"B{r}", title, bold=True, size=11, fill="F9FAFB")
        ws.merge_cells(f"C{r}:F{r}")
        setc(ws, f"C{r}", desc, size=10, fill="F9FAFB", color="374151")
        ws.row_dimensions[r].height = 28
        r += 1

    r += 1
    section_header(ws, r, "Re-running the decoder", color="6B7280", height=20)
    r += 1
    notes = [
        "Anytime your DIM share URLs change, re-run:    python3 decode_dim.py",
        "To add a new loadout without editing JSON:     python3 add_loadout.py",
        "Only the DIM LOADOUTS (FULL) tab is rewritten — every other tab is preserved.",
    ]
    for n in notes:
        ws.merge_cells(f"A{r}:F{r}")
        setc(ws, f"A{r}", f"  • {n}", italic=True, size=10, color="4B5563")
        r += 1

    r += 1
    section_header(ws, r, "Need help?", color="1F2937", height=20)
    r += 1
    ws.merge_cells(f"A{r}:F{r}")
    setc(ws, f"A{r}", "  See README.md and docs/ in the repo for screenshots and step-by-step setup.",
         italic=True, size=10, color="4B5563")


def build_priorities(wb):
    ws = wb.create_sheet("PRIORITIES")
    widths = [4, 28, 18, 22, 12, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    section_header(ws, 1, "PRIORITIES — What to Chase Next",
                   color="111827", height=28)
    ws["A1"].font = Font(name="Arial", bold=True, size=16, color="FFFFFF")

    r = 3
    # 1. Exotics
    section_header(ws, r, "1. TOP PRIORITY EXOTICS (ranked)", color="F59E0B", height=22)
    r += 1
    cols = ["#", "Exotic", "Slot", "Source", "Time", "Why / Unlocks"]
    for i, h in enumerate(cols, 1):
        setc(ws, f"{get_column_letter(i)}{r}", h, bold=True, fill="F3F4F6",
             align="center" if h == "#" else "left")
    r += 1
    sample_rows = [
        (1, "Example: Mataiodoxia", "Chest (Warlock)",
         "Edge of Fate exotic mission", "2-4h",
         "Replace with your own exotic chase. Rank them 1-10."),
        (2, "", "", "", "", ""),
        (3, "", "", "", "", ""),
    ]
    for row in sample_rows:
        for i, v in enumerate(row, 1):
            setc(ws, f"{get_column_letter(i)}{r}", v,
                 align="center" if i == 1 else "left",
                 italic=(r == 5 and i > 1),
                 color="9CA3AF" if r == 5 else "000000",
                 size=10)
        r += 1
    r += 1

    # 2. Campaigns
    section_header(ws, r, "2. CAMPAIGNS / STORY (in order)", color="F59E0B", height=22)
    r += 1
    cols2 = ["#", "Campaign / Quest", "Status", "Unlocks", "Time", "Notes"]
    for i, h in enumerate(cols2, 1):
        setc(ws, f"{get_column_letter(i)}{r}", h, bold=True, fill="F3F4F6",
             align="center" if h == "#" else "left")
    r += 1
    for i in range(1, 4):
        setc(ws, f"A{r}", i, align="center")
        for c in range(2, 7):
            setc(ws, f"{get_column_letter(c)}{r}", "", size=10, fill="FAFAFA")
        r += 1
    r += 1

    # 3. Farming
    section_header(ws, r, "3. OPTIMAL FARMING ROUTES (Tier 5 / A+ drops)",
                   color="F59E0B", height=22)
    r += 1
    cols3 = ["#", "Mission", "Type", "Best For", "Run Time", "Strategy"]
    for i, h in enumerate(cols3, 1):
        setc(ws, f"{get_column_letter(i)}{r}", h, bold=True, fill="F3F4F6",
             align="center" if h == "#" else "left")
    r += 1
    for i in range(1, 5):
        setc(ws, f"A{r}", i, align="center")
        for c in range(2, 7):
            setc(ws, f"{get_column_letter(c)}{r}", "", size=10, fill="FAFAFA")
        r += 1
    r += 1

    # 4. Mechanics
    section_header(ws, r, "4. GRINDING MECHANICS (how to hit A+ / Tier 5)",
                   color="F59E0B", height=22)
    r += 1
    cols4 = ["", "Mechanic", "", "Effect", "", "How to Use"]
    for i, h in enumerate(cols4, 1):
        setc(ws, f"{get_column_letter(i)}{r}", h, bold=True, fill="F3F4F6")
    r += 1
    for _ in range(4):
        for c in range(1, 7):
            setc(ws, f"{get_column_letter(c)}{r}", "", size=10, fill="FAFAFA")
        r += 1
    r += 1

    # 5. Mods
    section_header(ws, r, "5. MODS — Current artifact + build (update each season)",
                   color="F59E0B", height=22)
    r += 1
    ws.merge_cells(f"A{r}:F{r}")
    setc(ws, f"A{r}",
         "  Drop your current artifact mod selections + armor mod loadout here. "
         "Track which artifact columns are unlocked, which mods you're running on armor, "
         "and your current build's stats / power.",
         italic=True, color="6B7280", size=10, fill="F9FAFB")
    ws.row_dimensions[r].height = 36


def build_wishlist(wb):
    ws = wb.create_sheet("WISHLIST")
    widths = [4, 28, 12, 14, 32, 32, 14, 32]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    section_header(ws, 1, "WISHLIST — weapons + armor you're hunting",
                   color="111827", height=28, cols=8)
    ws["A1"].font = Font(name="Arial", bold=True, size=16, color="FFFFFF")

    r = 3
    ws.merge_cells(f"A{r}:H{r}")
    setc(ws, f"A{r}",
         "  Column order matches DIM's CSV export. You can paste rows directly "
         "from DIM → Settings → Export CSV (Weapons / Armor).",
         italic=True, color="4B5563", size=10, fill="F9FAFB")
    ws.row_dimensions[r].height = 22
    r += 2

    headers = ["#", "Name", "Tier", "Type", "Element / Subclass",
               "Source", "Owned?", "Notes / Desired Roll"]
    for i, h in enumerate(headers, 1):
        setc(ws, f"{get_column_letter(i)}{r}", h, bold=True, fill="F3F4F6",
             align="center" if h in ("#", "Owned?") else "left")
    r += 1

    sample = [
        (1, "Example: Sullen Claw", "Exotic", "Sword", "Void",
         "Equilibrium dungeon", "No", "Black blade — Sage + Swordmaster 2pc."),
        (2, "Example: Conditional Finality", "Exotic", "Shotgun", "Stasis + Solar",
         "Root of Nightmares", "No", "Freeze + scorch shotgun."),
    ]
    for row in sample:
        for i, v in enumerate(row, 1):
            setc(ws, f"{get_column_letter(i)}{r}", v,
                 align="center" if i in (1, 7) else "left",
                 italic=True, color="9CA3AF", size=10)
        r += 1
    for _ in range(12):
        setc(ws, f"A{r}", "", align="center", size=10, fill="FAFAFA")
        for c in range(2, 9):
            setc(ws, f"{get_column_letter(c)}{r}", "", size=10, fill="FAFAFA")
        r += 1


def build_class_sheet(wb, class_name):
    sheet_name = f"{class_name.upper()} BUILDS"
    ws = wb.create_sheet(sheet_name)
    widths = [4, 24, 16, 20, 22, 22, 22, 36]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    color = CLASS_COLORS[class_name]
    section_header(ws, 1, f"{class_name.upper()} BUILDS",
                   color=color, height=28, cols=8)
    ws["A1"].font = Font(name="Arial", bold=True, size=16, color="FFFFFF")

    r = 3
    ws.merge_cells(f"A{r}:H{r}")
    setc(ws, f"A{r}",
         f"  One row per {class_name} build. Track super, exotic armor, weapons, "
         "mods, and what activity each build is for.",
         italic=True, color="4B5563", size=10, fill="F9FAFB")
    ws.row_dimensions[r].height = 22
    r += 2

    headers = ["#", "Build Name", "Subclass", "Exotic Armor",
               "Kinetic", "Energy", "Heavy", "Notes / Activity"]
    for i, h in enumerate(headers, 1):
        setc(ws, f"{get_column_letter(i)}{r}", h, bold=True, fill="F3F4F6",
             align="center" if h == "#" else "left")
    r += 1

    # Seed 2-3 example rows so users see column meaning
    seeds = CLASS_EXAMPLES.get(class_name, [])
    for i, seed in enumerate(seeds, 1):
        setc(ws, f"A{r}", i, align="center", italic=True, color="9CA3AF")
        for c, val in enumerate(seed, 2):
            setc(ws, f"{get_column_letter(c)}{r}", val,
                 size=10, italic=True, color="9CA3AF")
        r += 1
    # Empty rows for the user to fill in
    for i in range(len(seeds) + 1, len(seeds) + 8):
        setc(ws, f"A{r}", i, align="center")
        for c in range(2, 9):
            setc(ws, f"{get_column_letter(c)}{r}", "", size=10, fill="FAFAFA")
        r += 1


def build_exotic_missions(wb):
    """Tracker for exotic-acquiring missions — current rotation + legacy farms."""
    ws = wb.create_sheet("EXOTIC MISSIONS")
    widths = [4, 28, 22, 18, 14, 12, 44]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    section_header(ws, 1, "EXOTIC MISSIONS — rotating + legacy",
                   color="111827", height=28, cols=7)
    ws["A1"].font = Font(name="Arial", bold=True, size=16, color="FFFFFF")

    r = 3
    ws.merge_cells(f"A{r}:G{r}")
    setc(ws, f"A{r}",
         "  Track which exotic missions you still need to run. Edit Status as you go. "
         "Use Priority (1 = highest) to sort. Re-runs for catalysts / rolls are fair game.",
         italic=True, color="4B5563", size=10, fill="F9FAFB")
    ws.row_dimensions[r].height = 28
    r += 2

    headers = ["#", "Mission", "Exotic Reward", "Source", "Status", "Priority", "Steps / Notes"]
    for i, h in enumerate(headers, 1):
        setc(ws, f"{get_column_letter(i)}{r}", h, bold=True, fill="F3F4F6",
             align="center" if h in ("#", "Status", "Priority") else "left")
    r += 1

    # Seed with currently rotating + recent exotic missions
    missions = [
        ('"Encore" (Edge of Fate)', "Mataiodoxia (Warlock chest)",
         "Edge of Fate campaign exit", "TODO", 1,
         "Complete Edge of Fate campaign → exotic mission unlocks. Replay for stat rolls."),
        ('"Fire & Ice" (Renegades)', "Praxic Blade + catalyst",
         "Renegades exotic mission", "TODO", 2,
         "First clear = Praxic Blade. Replay for catalyst (Upper Hand / Hyperblade)."),
        ('"The Way Between" quest', "Heirloom catalyst entry",
         "Praxic Temple post-Equilibrium", "TODO", 3,
         "Insert Integrated Relic at Praxic Temple after 1 Equilibrium clear."),
        ('"Operation: Seraph\'s Shield"', "Revision Zero + catalysts",
         "Seraph Bunker (legacy, rotating)", "Optional", 5,
         "4 catalysts: Hakke Heavy Burst, Outlaw, Pugilist, Killing Wind. Solo doable."),
        ('"Node.Ovrd.AVALON"', "Vexcalibur glaive",
         "Neomuna Vex network (legacy)", "Optional", 6,
         "Free void glaive with unique perk. Run weekly for upgrades to its perk."),
        ('"Presage"', "Dead Man's Tale (DMT)",
         "Cabal ship via Tangled Shore (legacy)", "Optional", 7,
         "Solo flawless triumph for catalyst. Hand cannon-feel scout, top-tier roll."),
        ('"Vox Obscura"', "Dead Messenger",
         "Mars (legacy, rotating)", "Optional", 8,
         "Triple-burst wave-frame GL. Excellent for ad clear and matching surge."),
        ('"Whisper of the Worm"', "Whisper of the Worm",
         "IO secret mission (legacy)", "Optional", 9,
         "Classic sniper. Worth running for catalyst and nostalgia."),
        ('"Outbreak Perfected"', "Outbreak Perfected",
         "Zero Hour (legacy)", "Optional", 10,
         "Nanite-spreading pulse. Legendary status for solo flawless."),
        ('"Starcrossed"', "Wish-Keeper",
         "Dreaming City (current rotation)", "Optional", 11,
         "Strand bow. Suspends targets. Great Hunter / Warlock CC."),
    ]
    for i, (name, exotic, src, status, priority, steps) in enumerate(missions, 1):
        setc(ws, f"A{r}", i, align="center", italic=True, color="9CA3AF")
        setc(ws, f"B{r}", name, italic=True, color="9CA3AF", size=10)
        setc(ws, f"C{r}", exotic, italic=True, color="9CA3AF", size=10)
        setc(ws, f"D{r}", src, italic=True, color="9CA3AF", size=10)
        setc(ws, f"E{r}", status, align="center", italic=True, color="9CA3AF", size=10)
        setc(ws, f"F{r}", priority, align="center", italic=True, color="9CA3AF", size=10)
        setc(ws, f"G{r}", steps, italic=True, color="9CA3AF", size=10)
        r += 1

    # Empty rows
    for i in range(len(missions) + 1, len(missions) + 6):
        setc(ws, f"A{r}", i, align="center")
        for c in range(2, 8):
            setc(ws, f"{get_column_letter(c)}{r}", "", size=10, fill="FAFAFA")
        r += 1


def build_mod_reference(wb, user_cfg=None):
    """
    Build the MOD REFERENCE sheet. If user_cfg is supplied (from user_config.json),
    use it to seed recommended mods from mod_recommender based on archetype + goals.
    """
    ws = wb.create_sheet("MOD REFERENCE")
    widths = [22, 22, 22, 22, 22, 36]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    section_header(ws, 1, "MOD REFERENCE — current season's artifact + armor mods",
                   color="111827", height=28)
    ws["A1"].font = Font(name="Arial", bold=True, size=16, color="FFFFFF")

    r = 3

    # If we have user config, drop a "your build focus" summary at the top
    if user_cfg and "build_focus" in user_cfg:
        bf = user_cfg["build_focus"]
        section_header(ws, r, "Your build focus  (from setup)",
                       color="7C3AED", height=22)
        r += 1
        ws.merge_cells(f"A{r}:F{r}")
        setc(ws, f"A{r}",
             f"  Class: {user_cfg.get('primary_class','?')}   ·   "
             f"Archetype: {bf.get('archetype','?')}   ·   "
             f"Goals: {', '.join(bf.get('goals', []))}",
             bold=True, size=11, color="FFFFFF", fill="7C3AED")
        ws.row_dimensions[r].height = 22
        r += 2

        # Recommended mods per slot
        try:
            from mod_recommender import recommend
            rec = recommend(bf.get("goals", []), bf.get("archetype", "Grenadier"),
                            user_cfg.get("primary_class", "Warlock"))
        except Exception:
            rec = None

        if rec:
            section_header(ws, r, "Recommended armor mods  (priority order — pick top N that fit energy)",
                           color="1F2937", height=22)
            r += 1
            for i, h in enumerate(["Slot", "Priority 1", "Priority 2", "Priority 3", "Priority 4+", "Notes"], 1):
                setc(ws, f"{get_column_letter(i)}{r}", h, bold=True, fill="F3F4F6")
            r += 1
            for slot_key, slot_label in [
                ("helmet", "Helmet"),
                ("gauntlets", "Gauntlets"),
                ("chest", "Chest"),
                ("legs", "Legs"),
                ("class_item", "Class Item"),
            ]:
                setc(ws, f"A{r}", slot_label, bold=True)
                mods = rec["mods"].get(slot_key, [])
                for i, mod in enumerate(mods[:4], 2):
                    setc(ws, f"{get_column_letter(i)}{r}", mod, size=10)
                if len(mods) > 4:
                    setc(ws, f"E{r}", " · ".join(mods[4:]), size=9, italic=True, color="6B7280")
                else:
                    setc(ws, f"E{r}", "", size=10, fill="FAFAFA")
                setc(ws, f"F{r}", "", size=10, fill="FAFAFA")
                r += 1
            r += 1
            ws.merge_cells(f"A{r}:F{r}")
            setc(ws, f"A{r}",
                 f"  Class tip ({user_cfg.get('primary_class', '?')} + {rec['primary_stat']} focus):  "
                 f"{rec['class_hint']}",
                 italic=True, size=10, color="4B5563", fill="F9FAFB", wrap=True)
            ws.row_dimensions[r].height = 32
            r += 2

    section_header(ws, r, "Artifact perks (5 columns × 5 tiers — current season)",
                   color="1F2937", height=22)
    r += 1
    for i in range(1, 6):
        setc(ws, f"{get_column_letter(i)}{r}", f"Column {i}",
             bold=True, fill="F3F4F6", align="center")
    setc(ws, f"F{r}", "Notes", bold=True, fill="F3F4F6", align="center")
    r += 1
    for _ in range(5):
        for c in range(1, 7):
            setc(ws, f"{get_column_letter(c)}{r}", "", size=10, fill="FAFAFA")
        r += 1
    r += 1

    section_header(ws, r, "Current build — stats + power",
                   color="1F2937", height=22)
    r += 1
    headers = ["Stat", "Value", "Tier"]
    for i, h in enumerate(headers, 1):
        setc(ws, f"{get_column_letter(i)}{r}", h, bold=True, fill="F3F4F6",
             align="left" if h == "Stat" else "center")
    ws.merge_cells(f"D{r}:F{r}")
    setc(ws, f"D{r}", "Read / Why", bold=True, fill="F3F4F6")
    r += 1
    for stat in ["Health", "Melee", "Grenade", "Super", "Class", "Weapons"]:
        setc(ws, f"A{r}", stat)
        setc(ws, f"B{r}", "", align="center", size=10, fill="FAFAFA")
        setc(ws, f"C{r}", "", align="center", size=10, fill="FAFAFA")
        ws.merge_cells(f"D{r}:F{r}")
        setc(ws, f"D{r}", "", size=10, fill="FAFAFA")
        r += 1
    r += 1

    section_header(ws, r, "Per-piece armor energy", color="1F2937", height=22)
    r += 1
    slots = ["Helmet", "Gauntlets", "Chest", "Legs", "Class Item", "Total"]
    for i, s in enumerate(slots, 1):
        setc(ws, f"{get_column_letter(i)}{r}", s, bold=True, align="center",
             fill="F3F4F6")
    r += 1
    for i in range(1, 7):
        setc(ws, f"{get_column_letter(i)}{r}", "", align="center", size=10,
             fill="FAFAFA")


def build_dim_loadouts_placeholder(wb):
    ws = wb.create_sheet("DIM LOADOUTS (FULL)")
    widths = [4, 14, 12, 36, 36, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    section_header(ws, 1, "DIM LOADOUTS (FULL) — run decode_dim.py to populate",
                   color="1F2937", height=28)
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")

    r = 3
    ws.merge_cells(f"A{r}:F{r}")
    setc(ws, f"A{r}",
         "  This sheet is auto-generated. Run `python3 decode_dim.py` in your terminal "
         "and it will fetch every DIM share URL listed in user_config.json, resolve every "
         "item / mod / aspect / fragment hash via the Bungie manifest, and rewrite this tab.",
         italic=True, color="4B5563", size=11, fill="F9FAFB")
    ws.row_dimensions[r].height = 60


def build_workbook(output_path, user_cfg=None):
    wb = Workbook()
    # Remove the default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    build_start_here(wb)
    build_priorities(wb)
    build_exotic_missions(wb)
    build_wishlist(wb)
    build_class_sheet(wb, "Hunter")
    build_class_sheet(wb, "Titan")
    build_class_sheet(wb, "Warlock")
    build_mod_reference(wb, user_cfg=user_cfg)
    build_dim_loadouts_placeholder(wb)

    wb.save(output_path)
    return output_path


def main():
    user_cfg = None
    cfg_path = Path("user_config.json")
    if cfg_path.exists():
        try:
            user_cfg = json.loads(cfg_path.read_text())
        except Exception:
            user_cfg = None

    if len(sys.argv) > 1:
        out = Path(sys.argv[1])
    elif user_cfg and "workbook_path" in user_cfg:
        out = Path(user_cfg["workbook_path"])
    else:
        out = Path("my_loadouts.xlsx")

    if out.exists():
        ans = input(f"{out} already exists. Overwrite? [y/N] ").strip().lower()
        if ans != "y":
            print("Aborted.")
            sys.exit(0)

    build_workbook(out, user_cfg=user_cfg)
    print(f"Wrote fresh workbook: {out}")


if __name__ == "__main__":
    main()
